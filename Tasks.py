import json
import os
import sys
import threading
import queue
import tkinter as tk
from tkinter import ttk, scrolledtext
from datetime import datetime
from time import sleep
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import playwright
from playwright.sync_api import sync_playwright, Playwright, TimeoutError
import warnings
import pyxlsb
import csv
import xlwings as xw
from playwright.sync_api import Page, Browser

base_path = os.getcwd()
bases_folder = os.path.join(base_path, "Bases")


warnings.filterwarnings("ignore", category=UserWarning)


def Process_A14_options(file_path, q):
    q.put(("status", "🔄 Inicializando o processamento dos arquivos..."))

    # Step 1: Load data (This part is unchanged)
    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext in [".xlsx", ".xlsm"]:
            df = pd.read_excel(file_path, engine="openpyxl", dtype=object)
        elif ext == ".xls":
            df = pd.read_excel(file_path, engine="xlrd", dtype=object)
        elif ext == ".xlsb":
            df = pd.read_excel(file_path, engine="pyxlsb", dtype=object)
        elif ext == ".csv":
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                sample = f.read(4096)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
                    delimiter = dialect.delimiter
                except csv.Error:
                    delimiter = ";" if ";" in sample else ","
            try:
                df = pd.read_csv(file_path, delimiter=delimiter, encoding="utf-8", engine="python", dtype=object)
            except Exception:
                df = pd.read_csv(file_path, delimiter=delimiter, encoding="latin-1", engine="python", dtype=object)
        else:
            q.put(("status", f"❌ Formato de arquivo não suportado: {ext}"))
            return
    except Exception as e:
        q.put(("status", f"❌ Erro ao ler arquivo: {e}"))
        return

    q.put(("status", f"✅ Arquivo carregado ({len(df)} linhas, {len(df.columns)} colunas)"))

    if 'CODICE_FAMIGLIA' not in df.columns:
        q.put(("status", f"❌ Coluna 'CODICE_FAMIGLIA' não encontrada. Colunas disponíveis: {list(df.columns)}"))
        return

    # Filter for 'PKG' rows
    df['CODICE_FAMIGLIA'] = df['CODICE_FAMIGLIA'].astype(str).str.strip().str.upper()
    df_pkg = df[df['CODICE_FAMIGLIA'] == 'PKG'].copy()

    if df_pkg.empty:
        q.put(("status", "⚠️ Nenhuma linha encontrada com CODICE_FAMIGLIA = 'PKG'."))
        return

    # Step 1: Find all columns with 'CODICE_OPTIONAL' in the name
    optional_cols = [col for col in df_pkg.columns if 'CODICE_OPTIONAL' in col]
    if not optional_cols:
        q.put(("status", "⚠️ Nenhuma coluna contendo 'CODICE_OPTIONAL' encontrada."))
        return

    # Step 2: The first optional column is for the PACK, the rest for CONTEÚDO
    pack_col_name = optional_cols[0]
    conteudo_cols = optional_cols[1:]
    
    q.put(("status", f"✅ Coluna do PACK: '{pack_col_name}'"))
    q.put(("status", f"✅ Colunas do CONTEÚDO: {len(conteudo_cols)} colunas"))

    processed_data = []
    for _, row in df_pkg.iterrows():
        # Step 3: Get the PACK value from the first optional column
        # Preserve exact string representation - no conversion whatsoever
        pack_value = row[pack_col_name]
        if pd.notna(pack_value):
            # Convert to string but preserve the original representation
            pack = str(pack_value).strip() if str(pack_value).strip().lower() != 'nan' else ""
        else:
            pack = ""
        
        # Step 4: Get and join all CONTEÚDO values from the other optional columns
        conteudo_values = []
        for col in conteudo_cols:
            value = row[col]
            if pd.notna(value):
                str_value = str(value).strip()
                # Exclude NaN strings
                if str_value and str_value.lower() != 'nan':
                    conteudo_values.append(str_value)
        
        conteudo = "*" + "*".join(conteudo_values) + "*" if conteudo_values else ""
        
        processed_data.append({'PACK': pack, 'CONTEÚDO': conteudo})

    df_result = pd.DataFrame(processed_data, columns=['PACK', 'CONTEÚDO'])
    q.put(("status", f"📦 {len(df_result)} registros prontos para atualização."))

    Base_folder = os.path.join(base_path, "Bases")

    if not os.path.exists(Base_folder):
        q.put(("status", f"❌ Pasta 'Bases' não encontrada: {Base_folder}"))
        return

    for filename in os.listdir(Base_folder):
        if 'BASE' in filename.upper() and not filename.startswith("~") and filename.lower().endswith(('.xlsb', '.xlsx', '.xlsm')):
            file_full_path = os.path.join(Base_folder, filename)
            q.put(("status", f"📁 Atualizando arquivo: {filename}"))
            try:
                app = xw.App(visible=False, add_book=False)
                app.display_alerts = False
                app.screen_updating = False

                wb = app.books.open(file_full_path, update_links=False)
                if 'A14' in [s.name for s in wb.sheets]:
                    ws = wb.sheets['A14']
                else:
                    ws = wb.sheets.add('A14')
                ws.clear_contents()
                
                # Write headers
                ws.range('A1').value = ['PACK', 'CONTEÚDO']
                
                # Format columns as text to preserve string format
                ws.range('A:B').number_format = '@'
                
                # Write data
                ws.range('A2').value = df_result.values.tolist()
                
                ws.autofit()
                wb.save()
                wb.close()
                # app.quit()
                q.put(("status", f"✅ Planilha 'A14' atualizada em {filename}"))
            except Exception as e:
                q.put(("status", f"❌ Falha ao processar {filename}: {e}"))

    q.put(("status", "🎉 Processamento concluído com sucesso."))


def download_A14(url_order,q,username,password,chromium_path, force=False) :
    # Check for recent A14 file if not forced
    if not force:
        os.makedirs("Dados", exist_ok=True)
        dados_dir = "Dados"
        today = datetime.now().date()
        recent_file = None
        max_date = None
        
        for filename in os.listdir(dados_dir):
            if filename.startswith("A14_") and filename.endswith(".xls"):
                try:
                    date_str = filename[4:-4]  # Remove "A14_" and ".xls"
                    file_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                    if max_date is None or file_date > max_date:
                        max_date = file_date
                        recent_file = filename
                except ValueError:
                    continue  # Skip malformed filenames
        
        if recent_file and max_date:
            days_old = (today - max_date).days
            if days_old <= 5:
                file_path = os.path.join(dados_dir, recent_file)
                q.put(("status", f"Arquivo A14 está atualizado (última atualização há {days_old} dias). Pulando download e processamento."))
                q.put(("progress", 65))
                return  # Skip download and processing
    
    with sync_playwright() as p:
        browser = None
        try :
            browser = p.chromium.launch(
                    headless=True, # Always headless for worker threads
                    executable_path=chromium_path,
                    # args=["--start-maximized"]
                )
            context = browser.new_context(viewport={'width': 1920, 'height': 1080})
            page = context.new_page()
            page.goto(url_order)

            q.put(("status", "Realizando login..."))
            q.put(("progress", 15))
        
            page.locator('[name="j_username"]').fill(username)
            page.locator('[name="j_password"]').fill(password)
            page.locator("button[type='submit']").click()
            q.put(("status", "Login realizado com sucesso!"))

            page.get_by_role("link", name="???tabstd???").hover()
        
            page.locator("li.ui-menuitem >> text=Download Table").click(timeout=200000)
            page.locator("[id=\"filter:codtab_label\"]").click()
            page.locator("[id=\"filter:codtab_panel\"]").get_by_text("???tabA14???").click()
            
            with page.expect_download() as download_info:
                page.get_by_role("button", name="Downloads").click()
            
            download = download_info.value

            today_str = datetime.now().date().isoformat()
            file_path = f"Dados/A14_{today_str}.xls"
            os.makedirs("Dados", exist_ok=True)
            
            if os.path.exists(file_path):
                os.remove(file_path)
            
            download.save_as(file_path)

            Process_A14_options(file_path,q)

            q.put(("status", f"Relatório A14 salvo como: {file_path}"))
            q.put(("status", "Downloads concluídos."))
            q.put(("progress", 65))
        except Exception as e:
            q.put(("status", f"ERRO ao processar o arquivo A14: {e}"))
        finally:
            pass


def download_por_modelo(url_oss, q, username, password, Modelos, chromium_path):
    """This function is the 'thread manager'. It creates and starts a thread for each model."""
    threads = []
    
    # Start all model threads
    for key, value in Modelos.items():
        if key == '611':
            q.put(("status", "Pulando modelo 611 com falha conhecida."))
            continue
            
        thread = threading.Thread(
            target=process_single_model,
            args=(url_oss, q, username, password, key, value, chromium_path)
        )
        threads.append(thread)
        thread.start()
        q.put(("status", f"Gerenciador: Thread iniciada para o modelo {key}."))

    # Wait for all model-processing threads to complete
    for thread in threads:
        thread.join()

    q.put(("status", "Gerenciador: Todas as threads de processamento de modelos foram concluídas."))

def process_single_model(url_oss: str, q: queue.Queue, username: str, password: str, key: str, value: str, chromium_path: str):
    """
    Processes a single model.
    This function is now COMPLETELY independent and thread-safe.
    Each thread downloads, updates BASE, and updates pivot tables.
    """
    # Each thread now creates its own Playwright instance and browser.
    with sync_playwright() as p:
        browser = None
        try:
            
            if chromium_path:
                browser = p.chromium.launch(
                headless=False,
                executable_path=chromium_path,
                args=[
                    "--start-maximized",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-infobars",
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                ]
                    )
                
            else : 
                browser = p.chromium.launch(
                    headless=False,
                    args=[
                        "--start-maximized",
                        "--disable-blink-features=AutomationControlled",
                        "--disable-infobars",
                        "--no-sandbox",
                        "--disable-dev-shm-usage",
                    ],
                )
            
            context = browser.new_context(
                no_viewport=True,  # Use full browser window size
            )
            
            
            page = context.new_page()

            q.put(("status", f"Modelo {key}: Iniciando thread..."))
            page.goto(url_oss, timeout=60000)

            page.locator('[name="USER_NAME"]').fill(username)
            page.locator('[name="PASSWORD"]').fill(password)
            page.locator(".signin").click()

            page.locator(".shellInstance").click()
            page.locator("#sequencer_ui_instances").select_option(value)
            page.get_by_role("link", name="Editor de programação").click(timeout=120000)

            q.put(("status", f"Modelo {key}: Aguardando página de relatório..."))

            frame = page.locator("iframe[name=\"appFrame\"]").content_frame
            inner_frame = frame.get_by_text("Your browser does not support").content_frame
            
            inner_frame.locator("#actionMenu").click(timeout=820000)

            with page.expect_download(timeout=120000) as download_info:
                inner_frame.get_by_text("Baixar CSV").click()
            
            download = download_info.value
            q.put(("status", f"Modelo {key}: Download iniciado."))

            os.makedirs("Dados", exist_ok=True)
            
            csv_path = os.path.join("Dados", f"{key}.csv")
            xlsx_path = os.path.join("Dados", f"{key}.xlsx")

            if os.path.exists(csv_path): os.remove(csv_path)
            if os.path.exists(xlsx_path): os.remove(xlsx_path)

            download.save_as(csv_path)
            q.put(("status", f"Modelo {key}: Salvo em {csv_path}"))
            
            # Close browser immediately after download completion
            try:
                context.close()
                browser.close()
                q.put(("status", f"Modelo {key}: Browser fechado após download."))
            except Exception as e:
                q.put(("status", f"Modelo {key}: Aviso ao fechar browser: {e}"))
            
            df = pd.read_csv(csv_path, low_memory=False)
            if "order_type" in df.columns and not df[df["order_type"] == "PRE"].empty:
                df_prev =df[df["order_type"] == "PRE"]
                df_prev.to_excel(xlsx_path, index=False, engine="xlsxwriter")

                q.put(("status", f"Modelo {key}: Excel criado em {xlsx_path}"))

                q.put(("status", f"Atualizando planilha Base para o Modelo {key}"))
                Atualizar_Base_Modelos(df_prev, key, q)
                
                q.put(("status", f"Planilha Base do Modelo {key}: Atualizado com sucesso "))
                
                # Update pivot tables using the shared GRIGLIA link
                q.put(("status", f"Modelo {key}: Atualizando tabelas dinâmicas..."))
                Atualizar_Links_Pivort_tables_Single_Model(key, q)
            
            q.put(("status", f"Modelo {key}: Processamento completo concluído."))

        except Exception as e:
            q.put(("status", f"ERRO na thread do modelo {key}: {e}"))
        finally:
            # Browser is now closed immediately after download; 'with sync_playwright()' handles any remaining cleanup.
            pass



def Atualizar_Base_Modelos(df_to_paste, model_key, q):
    """
    Finds the correct 'BASE' file, clears it, pastes the data (without the header),
    copies formatting from the second row, and autofills formulas.
    """
    q.put(("status", f"Modelo {model_key}: Procurando arquivo BASE para atualizar..."))
    
    try:
        # Step 1: Find the target file in the 'Bases' subfolder

        if not os.path.exists(bases_folder):
            q.put(("status", f"Modelo {model_key}: ERRO - Pasta 'Bases' não encontrada."))
            return

        target_file_path = None
        target_filename = None
        for filename in os.listdir(bases_folder):
            if (filename.upper().startswith('BASE') and 
                model_key in filename and 
                not filename.startswith("~")):
                target_file_path = os.path.join(bases_folder, filename)
                target_filename = filename
                break

        if not target_file_path:
            q.put(("status", f"Modelo {model_key}: ERRO - Nenhum arquivo BASE encontrado para este modelo."))
            return
            
        q.put(("status", f"Modelo {model_key}: Arquivo encontrado: {target_filename}"))

        # Use a new app instance for each thread to ensure stability
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        
        try:
            q.put(("status", f"Modelo {model_key}: Abrindo arquivo {target_filename}..."))
            wb = app.books.open(target_file_path, update_links=False)
            
            try:
                ws = wb.sheets['ARQUIVO PREVISÕES']
                q.put(("status", f"Modelo {model_key}: Limpando dados antigos da planilha..."))
                
                # Prepare the data: Remove only the first row (the header).
                df_data_only = df_to_paste.iloc[1:]
                
                # STOP AT AE TO AVOID PIVOT TABLE
                try:
                    used_range = ws.range('B3').expand('table')
                    if used_range.rows.count > 0:
                        last_used_row = used_range.last_cell.row
                        ws.range(f'B3:AE{last_used_row}').clear_contents()
                except:
                    # If no data exists, just clear a small range
                    ws.range('B3:AE1000').clear_contents()
                                   
                # Proceed only if there is data left after removing the header.
                if not df_data_only.empty:
                    # Calculate the last row based on data size
                    num_data_rows = len(df_data_only)
                    last_row = 2 + num_data_rows  # Row 2 is template, data starts at row 3
                    
                    q.put(("status", f"Modelo {model_key}: Colando dados ({num_data_rows} linhas)..."))
                    
                    # 1. PASTE DATA FIRST at B3
                    start_cell = ws.range('B3')
                    start_cell.options(header=False, index=False).value = df_data_only.values
                    
                    q.put(("status", f"Modelo {model_key}: Aplicando formatação e fórmulas..."))
                    
                    # 2. COPY FORMAT AND FORMULAS - more efficiently using copy/paste
                    template_range = ws.range('A2:AE2')
                    target_range = ws.range(f'A3:AE{last_row}')
                    
                    # Copy template formatting
                    template_range.api.Copy()
                    # Paste formats only (no values) to avoid overwriting data
                    target_range.api.PasteSpecial(Paste=-4122)  # xlPasteFormats
                    
                    # Copy formulas from column A (always has formulas)
                    ws.range('A2').copy(ws.range(f'A3:A{last_row}'))
                    
                    # Copy formulas from columns Z-AE (these columns have formulas)
                    ws.range('Z2:AE2').copy(ws.range(f'Z3:AE{last_row}'))
                    
                    # Copy formulas from columns after AE (if any exist)
                    template_formulas = ws.range('AF2:ZZ2').formula
                    if any(f for f in template_formulas if f and str(f).startswith('=')):
                        ws.range('AF2:ZZ2').copy(ws.range(f'AF3:ZZ{last_row}'))
                    
                    # Clear clipboard
                    app.api.CutCopyMode = False
                    
                    q.put(("status", f"Modelo {model_key}: Dados colados e formatados em '{target_filename}' ({num_data_rows} linhas)."))

                else:
                    # If there's no data after removing the header, log a warning.
                    q.put(("status", f"Modelo {model_key}: AVISO - Sem dados para colar. Planilha '{target_filename}' foi limpa."))
                    
                q.put(("status", f"Modelo {model_key}: Salvando arquivo {target_filename}..."))
                wb.save()
            except Exception as sheet_error:
                q.put(("status", f"Modelo {model_key}: ERRO ao processar a planilha em '{target_filename}': {sheet_error}"))
                raise
            finally:
                wb.close()
        finally:
            app.quit()

    except Exception as e:
        q.put(("status", f"Modelo {model_key}: ERRO FATAL em Atualizar_Base_Modelos: {e}"))
        
               
def Atualizar_Links_Pivort_tables_Single_Model(model_key, q):
    
    try:
        
        if not os.path.exists(bases_folder):
            q.put(("status", f"Modelo {model_key}: ERRO - Pasta 'Bases' não encontrada."))
            return
        
        # Find BASE file for this model
        base_file = None
        for filename in os.listdir(bases_folder):
            if (filename.upper().startswith('BASE') and 
                model_key in filename and 
                not filename.startswith("~")):
                base_file = filename
                break
        
        if not base_file:
            q.put(("status", f"Modelo {model_key}: ERRO - Arquivo BASE não encontrado."))
            return
        
        base_file_path = os.path.join(bases_folder, base_file)
        q.put(("status", f"Modelo {model_key}: Arquivo BASE encontrado: {base_file}"))
        
        # Open Excel app for this thread (each thread needs its own app instance)
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        
        try:
            # First, open GRIGLIA file in this Excel instance so it's accessible
            griglia_file = None
            for filename in os.listdir(bases_folder):
                if 'GRIGLIA OPCIONAIS' in filename.upper() and not filename.startswith("~"):
                    griglia_file = filename
                    break
            
            if griglia_file:
                griglia_path = os.path.join(bases_folder, griglia_file)
                wb_griglia_local = app.books.open(griglia_path, update_links=False)
                q.put(("status", f"Modelo {model_key}: GRIGLIA aberto na mesma instância do Excel."))
            
            # Open BASE file for this model
            wb_base = app.books.open(base_file_path, update_links=False)
            
            try:
                # Access ANALYSIS sheet
                if 'ANALYSIS' not in [s.name for s in wb_base.sheets]:
                    q.put(("status", f"Modelo {model_key}: ERRO - Planilha 'ANALYSIS' não encontrada."))
                    return
                
                ws_analysis = wb_base.sheets['ANALYSIS']
                q.put(("status", f"Modelo {model_key}: Planilha 'ANALYSIS' acessada."))
                
                # Find and update pivot tables
                pivot_tables = ws_analysis.api.PivotTables()
                
                if pivot_tables.Count == 0:
                    q.put(("status", f"Modelo {model_key}: AVISO - Nenhuma tabela dinâmica encontrada."))
                else:
                    # Find last valid row in GRIGLIA BANCO sheet (column A)
                    try:
                        ws_griglia_banco = wb_griglia_local.sheets['BANCO']
                        # Find last row with data in column A starting from A1
                        last_row_griglia = ws_griglia_banco.range('A1').end('down').row
                        q.put(("status", f"Modelo {model_key}: Última linha válida em GRIGLIA BANCO: {last_row_griglia}"))
                    except Exception as e:
                        q.put(("status", f"Modelo {model_key}: ERRO ao encontrar última linha em GRIGLIA: {e}. Usando linha padrão 3207."))
                        last_row_griglia = 3207
                    
                    for i in range(1, pivot_tables.Count + 1):
                        pivot_table = pivot_tables.Item(i)
                        pivot_name = pivot_table.Name
                        
                        # Check pivot table location
                        pivot_location = pivot_table.TableRange1.Address
                        q.put(("status", f"Modelo {model_key}: Processando '{pivot_name}' em {pivot_location}"))
                        
                        # Update the data source using the shared GRIGLIA link
                        try:
                            wb_name = os.path.basename(griglia_path)  # e.g. "GRIGLIA OPCIONAIS 01.12.2025.xlsb"
                            source_data = f"'[{wb_name}]BANCO'!$A$1:$Q${last_row_griglia}"

                            q.put(("status", f"Modelo {model_key}: Usando fonte de dados: {source_data}"))

                            # --- Use PivotCaches.Create to force a fresh cache (xlDatabase = 1) ---
                            # Excel constant: xlDatabase = 1
                            xlDatabase = 1

                            # Create a new pivot cache on the BASE workbook, pointing to the external range
                            new_cache = wb_base.api.PivotCaches().Create(xlDatabase, source_data)

                            try:
                                pivot_table.ChangePivotCache(new_cache)
                            except Exception as inner_e:
                                # Some Excel versions prefer assignment, but ChangePivotCache is standard
                                try:
                                    pivot_table.PivotCache = new_cache
                                except Exception:
                                    # fallback: raise the original inner exception to be caught below
                                    raise inner_e

                            # Refresh the cache (and the pivot)
                            try:
                                new_cache.Refresh()
                            except Exception:
                                # If cache refresh fails, still try to refresh the table
                                pass

                            # small pause to let Excel update internal structures
                            sleep(1)

                            q.put(("status", f"Modelo {model_key}: Atualizando tabela dinâmica..."))
                            pivot_table.RefreshTable()

                            q.put(("status", f"Modelo {model_key}: '{pivot_name}' atualizada com sucesso."))
                        except Exception as e:
                            # If we still fail, print header row from GRIGLIA BANCO for debugging
                            try:
                                wb_name = os.path.basename(griglia_path)
                                # wb_griglia_local should exist in locals() since you opened it earlier
                                if 'wb_griglia_local' in locals():
                                    try:
                                        headers = wb_griglia_local.sheets['BANCO'].range('A1:Q1').value
                                    except Exception:
                                        # sometimes COM returns nested tuples for a single-row range; normalize
                                        try:
                                            headers = wb_griglia_local.sheets['BANCO'].range('A1:Q1').options(ndim=1).value
                                        except Exception:
                                            headers = None
                                else:
                                    headers = None
                            except Exception:
                                headers = None

                            q.put(("status", f"Modelo {model_key}: ERRO ao atualizar '{pivot_name}': {e}"))
                            if headers is not None:
                                # show each header with its index and a repr() to reveal spaces/unprintables
                                header_debug = []
                                for idx, h in enumerate(headers, start=1):
                                    header_debug.append(f"col {idx}: {repr(h)}")
                                q.put(("status", f"Modelo {model_key}: Headers A1..Q1 -> " + " | ".join(header_debug)))
                            else:
                                q.put(("status", f"Modelo {model_key}: Não foi possível ler cabeçalhos para diagnóstico."))

                
                # Save the BASE file (GRIGLIA must stay open until after save)
                q.put(("status", f"Modelo {model_key}: Salvando arquivo BASE..."))
                wb_base.save()
                q.put(("status", f"Modelo {model_key}: Tabelas dinâmicas atualizadas e arquivo salvo."))
                
                # Update PREVISÕES X ISTOGRAMA file (keep BASE file open)
                Atualizar_Previsao_X_Istograma(model_key, q, app, wb_base, base_file)
                
                
            finally:
                wb_base.close()
            
            # Close local GRIGLIA file AFTER BASE file is closed
            if griglia_file and 'wb_griglia_local' in locals():
                wb_griglia_local.close()
                q.put(("status", f"Modelo {model_key}: GRIGLIA local fechado."))
                
        finally:
            app.quit()
        
    except Exception as e:
        q.put(("status", f"Modelo {model_key}: ERRO em Atualizar_Links_Pivort_tables_Single_Model: {e}"))
        import traceback
        q.put(("status", f"Modelo {model_key}: Traceback: {traceback.format_exc()}"))
        
           
def Atualizar_Previsao_X_Istograma(model_key, q, app, wb_base, base_filename):
    
    try:
        q.put(("status", f"Modelo {model_key}: Iniciando atualização PREVISÕES X ISTOGRAMA..."))
        
        if not os.path.exists(bases_folder):
            q.put(("status", f"Modelo {model_key}: ERRO - Pasta 'Bases' não encontrada."))
            return
        
        # Find PREVISÕES X ISTOGRAMA file for this model
        previsoes_file = None
        for filename in os.listdir(bases_folder):
            if (filename.upper().startswith('PREVISÕES')  and "ISTOGRAMA" in filename.upper() and 
                model_key in filename and 
                not filename.startswith("~")):
                previsoes_file = filename
                break
            
        if not previsoes_file:
            q.put(("status", f"Modelo {model_key}: AVISO - Arquivo PREVISÕES X ISTOGRAMA não encontrado. Pulando..."))
            return
        
        previsoes_file_path = os.path.join(bases_folder, previsoes_file)
        q.put(("status", f"Modelo {model_key}: Arquivo PREVISÕES encontrado: {previsoes_file}"))
        
        # Open PREVISÕES X ISTOGRAMA file in the same Excel instance
        wb_previsoes = app.books.open(previsoes_file_path, update_links=False)
        
        try:
            # Access ANÁLISE PREVISÕES OPCIONAIS or ANÁLISE PREVISÕES POR OPCIONAL sheet
            sheet_names = ['ANÁLISE PREVISÕES OPCIONAIS', 'ANÁLISE PREVISÕES POR OPCIONAL','ANÁLISE PREVISÕES']
            sheet_name = None
            
            for name in sheet_names:
                if name in [s.name for s in wb_previsoes.sheets]:
                    sheet_name = name
                    break
            
            if sheet_name is None:
                q.put(("status", f"Modelo {model_key}: ERRO - Nenhuma planilha encontrada com os nomes esperados: {sheet_names}"))
                return
            
            ws_previsoes = wb_previsoes.sheets[sheet_name]
            q.put(("status", f"Modelo {model_key}: Planilha '{sheet_name}' acessada."))
            
            # Find and update pivot tables
            pivot_tables = ws_previsoes.api.PivotTables()
            
            if pivot_tables.Count == 0:
                q.put(("status", f"Modelo {model_key}: AVISO - Nenhuma tabela dinâmica encontrada em '{sheet_name}'."))
            else:
                q.put(("status", f"Modelo {model_key}: Encontradas {pivot_tables.Count} tabela(s) dinâmica(s)."))
                
                # Find last valid row in BASE ANALYSIS sheet (column Q starting from Q5)
                try:
                    ws_base_analysis = wb_base.sheets['ANALYSIS']
                    # Find last row with data in column Q starting from Q5
                    last_row_analysis = ws_base_analysis.range('Q5').end('down').row
                    q.put(("status", f"Modelo {model_key}: Última linha válida em BASE ANALYSIS (coluna Q): {last_row_analysis}"))
                except Exception as e:
                    q.put(("status", f"Modelo {model_key}: ERRO ao encontrar última linha em ANALYSIS: {e}. Usando linha padrão 49."))
                    last_row_analysis = 49
                
                for i in range(1, pivot_tables.Count + 1):
                    pivot_table = pivot_tables.Item(i)
                    pivot_name = pivot_table.Name
                    
                    # Check pivot table location
                    pivot_location = pivot_table.TableRange1.Address
                    q.put(("status", f"Modelo {model_key}: Processando '{pivot_name}' em {pivot_location}"))
                    
                    # Update the data source using the BASE file ANALYSIS sheet range Q5:EO49
                    try:
                        # Build external link to BASE file's ANALYSIS sheet
                        source_data = f"'[{base_filename}]ANALYSIS'!$Q$5:$EO${last_row_analysis}"
                        
                        q.put(("status", f"Modelo {model_key}: Usando fonte de dados: {source_data}"))
                        
                        # Create a new pivot cache pointing to the BASE file's ANALYSIS range
                        xlDatabase = 1
                        new_cache = wb_previsoes.api.PivotCaches().Create(xlDatabase, source_data)
                        
                        try:
                            pivot_table.ChangePivotCache(new_cache)
                        except Exception as inner_e:
                            try:
                                pivot_table.PivotCache = new_cache
                            except Exception:
                                raise inner_e
                        
                        # Refresh the cache and pivot table
                        try:
                            new_cache.Refresh()
                        except Exception:
                            pass
                        
                        sleep(1)
                        
                        q.put(("status", f"Modelo {model_key}: Atualizando tabela dinâmica '{pivot_name}'..."))
                        pivot_table.RefreshTable()
                        
                        q.put(("status", f"Modelo {model_key}: '{pivot_name}' atualizada com sucesso."))
                        
                    except Exception as e:
                        q.put(("status", f"Modelo {model_key}: ERRO ao atualizar '{pivot_name}': {e}"))
                        import traceback
                        q.put(("status", f"Modelo {model_key}: Traceback: {traceback.format_exc()}"))
            
            # Save the PREVISÕES X ISTOGRAMA file
            q.put(("status", f"Modelo {model_key}: Salvando arquivo PREVISÕES X ISTOGRAMA..."))
            wb_previsoes.save()
            q.put(("status", f"Modelo {model_key}: PREVISÕES X ISTOGRAMA atualizado e salvo com sucesso."))
            
            
            Criar_Dados_A_Analizar_Previsoes(wb_previsoes, model_key, q)
            
        finally:
            wb_previsoes.close()
            q.put(("status", f"Modelo {model_key}: Arquivo PREVISÕES X ISTOGRAMA fechado."))
            
    except Exception as e:
        q.put(("status", f"Modelo {model_key}: ERRO em Atualizar_Previsao_X_Istograma: {e}"))
        import traceback
        q.put(("status", f"Modelo {model_key}: Traceback: {traceback.format_exc()}"))
    
    
        
        
def Criar_Dados_A_Analizar_Previsoes(wb_previsoes, model_key, q):
    """
    Extract rows with non-zero/non-empty values from the data range (DL9:DW+) 
    and copy them to a new sheet 'PREVISÕES A CORRIGIR' with formatting.
    """
    try:
        q.put(("status", f"Modelo {model_key}: Iniciando criação de dados a analisar..."))
        
        # Access the correct sheet
        sheet_names = ['ANÁLISE PREVISÕES OPCIONAIS', 'ANÁLISE PREVISÕES POR OPCIONAL']
        sheet_name = None
        
        for name in sheet_names:
            if name in [s.name for s in wb_previsoes.sheets]:
                sheet_name = name
                break
        
        if sheet_name is None:
            q.put(("status", f"Modelo {model_key}: ERRO - Nenhuma planilha encontrada para análise de previsões."))
            return
        
        ws_source = wb_previsoes.sheets[sheet_name]
        q.put(("status", f"Modelo {model_key}: Processando planilha '{sheet_name}'..."))
        
        # Step 1: Find the last column in row 6 starting from DL
        try:
            # DL is column 116 (D=4, L=12 -> 4*26 + 12 = 116)
            start_col = 116  # DL
            last_col_detected = ws_source.range(f'DL6').end('right').column
            # Subtract 1 to exclude the observations column
            last_col = last_col_detected - 1
            q.put(("status", f"Modelo {model_key}: Última coluna detectada: {get_column_letter(last_col_detected)}, usando {get_column_letter(last_col)} (excluindo observações)"))
        except Exception as e:
            q.put(("status", f"Modelo {model_key}: ERRO ao detectar última coluna: {e}. Usando DV como padrão."))
            last_col = 125  # DV = 4*26 + 22 = 125 (DW-1)
        
        # Step 2: Find the last row with data - use used range or column A
        try:
            # Try to find last row by checking the used range
            used_range = ws_source.used_range
            last_row = used_range.last_cell.row
            
            # If that gives a small number, try column A specifically
            if last_row < 50:
                # Start from A9 and go down to find last row with data
                test_range = ws_source.range('A9')
                if test_range.value is not None:
                    last_row = ws_source.range('A9').end('down').row
                else:
                    # If A9 is empty, use the entire column
                    last_row = ws_source.range('A1048576').end('up').row
            
            q.put(("status", f"Modelo {model_key}: Última linha com dados: {last_row}"))
        except Exception as e:
            q.put(("status", f"Modelo {model_key}: ERRO ao detectar última linha: {e}"))
            return
        
        # Step 3: Create or clear the target sheet
        target_sheet_name = 'PREVISÕES A CORRIGIR'
        
        if target_sheet_name in [s.name for s in wb_previsoes.sheets]:
            wb_previsoes.sheets[target_sheet_name].delete()
            q.put(("status", f"Modelo {model_key}: Planilha '{target_sheet_name}' deletada."))
        
        ws_target = wb_previsoes.sheets.add(target_sheet_name)
        q.put(("status", f"Modelo {model_key}: Planilha '{target_sheet_name}' criada."))
        
        # Step 4: Copy header rows (rows 4 to 6) - values with formatting
        try:
            q.put(("status", f"Modelo {model_key}: Copiando cabeçalho (linhas 4-6) com formatação..."))
            
            # Unhide all rows and columns in source to ensure complete copy
            try:
                ws_source.api.Rows.Hidden = False
                ws_source.api.Columns.Hidden = False
                q.put(("status", f"Modelo {model_key}: Linhas e colunas ocultas reveladas na origem."))
            except:
                pass
            
            # Copy the entire range with formatting using Windows API
            source_range = ws_source.range(f'A3:{get_column_letter(last_col)}6')
            target_range = ws_target.range('A1')
            
            # Use Windows API to copy values and formatting (no formulas)
            source_range.api.Copy()
            target_range.api.PasteSpecial(Paste=-4163)  # xlPasteFormats (formats only)
            target_range.api.PasteSpecial(Paste=-4163)  # xlPasteFormats second to ensure merge cells
            target_range.api.PasteSpecial(Paste=-4122)  # xlPasteValuesAndNumberFormats (values + number formats)
            wb_previsoes.app.api.CutCopyMode = False
            
            q.put(("status", f"Modelo {model_key}: Cabeçalho copiado (valores + formatação, sem fórmulas)."))
        except Exception as e:
            q.put(("status", f"Modelo {model_key}: ERRO ao copiar cabeçalho: {e}"))
        
        # Step 5: Filter and copy rows with non-zero/non-empty values
        rows_to_copy = []
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(last_col)
        
        q.put(("status", f"Modelo {model_key}: Analisando linhas de 9 até {last_row} nas colunas {start_col_letter} até {end_col_letter}..."))
        
        for row_num in range(9, last_row + 1):
            # Get values from DL to last_col for this row
            data_range = ws_source.range(f'{start_col_letter}{row_num}:{end_col_letter}{row_num}')
            values = data_range.value
            
            # Check if any value is non-zero and non-empty
            has_valid_data = False
            
            # Handle both single value and list of values
            if not isinstance(values, (list, tuple)):
                values = [values]
            else:
                # Flatten if it's a nested list (single row returns as list of lists)
                if len(values) > 0 and isinstance(values[0], (list, tuple)):
                    values = values[0]
            
            
            for val in values:
                if val is not None and val != "" and val != 0:
                    # Check if it's a numeric value different from 0
                    if isinstance(val, (int, float)) and val != 0:
                        has_valid_data = True
                        break
                    # Check if it's a non-empty string
                    elif isinstance(val, str) and val.strip() != "" and val.strip() != "0":
                        has_valid_data = True
                        break
            
            if has_valid_data:
                rows_to_copy.append(row_num)
        
        q.put(("status", f"Modelo {model_key}: Encontradas {len(rows_to_copy)} linhas com dados a corrigir."))
        
        # Step 6: Copy filtered rows with formatting
        if rows_to_copy:
            q.put(("status", f"Modelo {model_key}: Copiando {len(rows_to_copy)} linhas para '{target_sheet_name}'..."))
            target_row = 5  # Start pasting from row 6 (rows 1-3 have header, 4-5 are buffer)
            
            for idx, source_row in enumerate(rows_to_copy, 1):
                try:
                    # Copy entire row with formatting - using direct range copy
                    source_range = ws_source.range(f'A{source_row}:{get_column_letter(last_col)}{source_row}')
                    target_range = ws_target.range(f'A{target_row}:{get_column_letter(last_col)}{target_row}')
                    
                    # Copy with API
                    source_range.api.Copy()
                    target_range.api.PasteSpecial(Paste=-4104)  # xlPasteAllUsingSourceTheme
                    
                    target_row += 1
                    
                    # Log progress every 10 rows
                    if idx % 10 == 0:
                        q.put(("status", f"Modelo {model_key}: Copiadas {idx}/{len(rows_to_copy)} linhas..."))
                    
                    # Clear clipboard every 50 rows to prevent memory issues
                    if idx % 50 == 0:
                        wb_previsoes.app.api.CutCopyMode = False
                        
                except Exception as row_error:
                    q.put(("status", f"Modelo {model_key}: ERRO ao copiar linha {source_row}: {row_error}"))
            
            wb_previsoes.app.api.CutCopyMode = False
            q.put(("status", f"Modelo {model_key}: {len(rows_to_copy)} linhas copiadas para '{target_sheet_name}'."))
            
            # Verify data was copied
            try:
                test_value = ws_target.range('A4').value
                q.put(("status", f"Modelo {model_key}: Verificação - Valor em A4: {test_value}"))
            except:
                pass
        else:
            q.put(("status", f"Modelo {model_key}: Nenhuma linha com dados a corrigir encontrada."))
        
        # Step 7: Save the workbook
        q.put(("status", f"Modelo {model_key}: Salvando planilha '{target_sheet_name}'..."))
        wb_previsoes.save()
        q.put(("status", f"Modelo {model_key}: Planilha '{target_sheet_name}' criada e salva com sucesso."))
        
    except Exception as e:
        q.put(("status", f"Modelo {model_key}: ERRO em Criar_Dados_A_Analizar_Previsoes: {e}"))
        import traceback
        q.put(("status", f"Modelo {model_key}: Traceback: {traceback.format_exc()}"))    
        
        
        